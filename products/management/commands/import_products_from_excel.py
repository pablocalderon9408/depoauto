from __future__ import annotations

import os
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Tuple

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.conf import settings
from django.core.files import File

from openpyxl import load_workbook
from django.utils.text import slugify

from products.models import Category, Product, ProductVariant


class Command(BaseCommand):
    help = "Importa categorías y productos desde un archivo Excel (.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            dest="file_path",
            default="data.xlsx",
            help="Ruta al archivo .xlsx (por defecto: data.xlsx en la raíz del proyecto)",
        )
        parser.add_argument(
            "--sheet",
            dest="sheet_name",
            default=None,
            help="Nombre de la hoja a importar (por defecto, la activa)",
        )
        parser.add_argument(
            "--update",
            action="store_true",
            dest="do_update",
            help="Si existe un producto con el mismo SKU, actualizar sus datos",
        )
        parser.add_argument(
            "--images-dir",
            dest="images_dir",
            default=None,
            help="Directorio base donde residen las imágenes (ej.: FOTOS). "
                 "Estructura sugerida: <base>/<categoria>/<SKU>/*.jpg",
        )
        parser.add_argument(
            "--replace-images",
            action="store_true",
            dest="replace_images",
            help="Si se especifica, elimina imágenes existentes del producto/variante antes de cargar nuevas.",
        )
        parser.add_argument(
            "--limit-images",
            dest="limit_images",
            type=int,
            default=0,
            help="Límite de imágenes a cargar por SKU (0 = sin límite).",
        )

    def handle(self, *args, **options):
        file_path = options["file_path"]
        sheet_name = options["sheet_name"]
        do_update = options["do_update"]
        images_dir = options.get("images_dir")
        replace_images = options.get("replace_images", False)
        limit_images: int = options.get("limit_images", 0) or 0

        abs_path = file_path
        if not os.path.isabs(abs_path):
            abs_path = os.path.join(settings.BASE_DIR, file_path)
        if not os.path.exists(abs_path):
            raise CommandError(f"No se encontró el archivo: {abs_path}")

        images_base_dir: str | None = None
        if images_dir:
            images_base_dir = images_dir
            if not os.path.isabs(images_base_dir):
                images_base_dir = os.path.join(settings.BASE_DIR, images_base_dir)
            if not os.path.isdir(images_base_dir):
                raise CommandError(f"--images-dir apunta a un directorio inexistente: {images_base_dir}")

        wb = load_workbook(abs_path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        headers = self._read_headers(ws)
        if not headers:
            raise CommandError("La hoja no contiene encabezados válidos en la primera fila.")

        self.stdout.write(self.style.NOTICE(f"Encabezados detectados: {headers}"))
        mapped = self._map_columns(headers)
        self.stdout.write(self.style.NOTICE(f"Mapeo de columnas: {mapped}"))

        created_categories = 0
        created_products = 0
        updated_products = 0

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = self._row_to_dict(row, headers)
                try:
                    category_name = (data.get(mapped["category"]) or "").strip()
                    if not category_name:
                        self.stdout.write(self.style.WARNING(f"Fila {row_idx}: sin categoría; saltando"))
                        continue
                    category, cat_created = Category.objects.get_or_create(name=category_name)
                    if cat_created:
                        created_categories += 1

                    name = (data.get(mapped["name"]) or "").strip()
                    sku = (data.get(mapped["sku"]) or "").strip()
                    if not sku:
                        self.stdout.write(self.style.WARNING(f"Fila {row_idx}: sin SKU; saltando"))
                        continue
                    if not name:
                        name = sku

                    # Campos opcionales: usar mapeo si existe
                    desc_col = mapped.get("description")
                    description = ((data.get(desc_col) if desc_col else "") or "").strip()

                    price_col = mapped.get("price")
                    price = self._parse_decimal(data.get(price_col)) if price_col else None

                    stock_col = mapped.get("stock")
                    stock = self._parse_int(data.get(stock_col)) if stock_col else None

                    image_col = mapped.get("image_url")
                    image_url = ((data.get(image_col) if image_col else "") or "").strip()

                    product_defaults = {
                        "name": name,
                        "category": category,
                        "description": description,
                        "price": price if price is not None else Decimal("0"),
                        "stock": stock if stock is not None else 0,
                        "is_active": True,
                    }
                    if image_url:
                        product_defaults["image_url"] = image_url

                    # Crear o actualizar por SKU asegurando slug único
                    obj = Product.objects.filter(sku=sku).first()
                    if obj is None:
                        unique_slug = self._make_unique_product_slug(name, sku)
                        payload = {**product_defaults, "sku": sku, "slug": unique_slug}
                        obj = Product(**payload)
                        obj.save()
                        created_products += 1
                    else:
                        if do_update:
                            for field, value in product_defaults.items():
                                setattr(obj, field, value)
                            obj.slug = self._make_unique_product_slug(name, sku, exclude_id=obj.id)
                            obj.save()
                            updated_products += 1

                    # Asegurar al menos una variante vacía
                    ProductVariant.objects.get_or_create(product=obj, name="")

                    # Carga de imágenes desde disco si se especificó --images-dir
                    if images_base_dir:
                        try:
                            loaded = self._load_images_for_product(
                                product=obj,
                                images_base_dir=images_base_dir,
                                category_name=category.name,
                                sku=sku,
                                replace_existing=replace_images,
                                limit=limit_images,
                            )
                            if loaded:
                                self.stdout.write(self.style.SUCCESS(
                                    f"Imágenes cargadas para SKU {sku}: {loaded} archivo(s)"))
                            else:
                                self.stdout.write(self.style.WARNING(
                                    f"Sin imágenes encontradas para SKU {sku}"))
                        except Exception as img_exc:
                            raise CommandError(f"Error cargando imágenes para SKU {sku}: {img_exc}") from img_exc
                except Exception as exc:
                    raise CommandError(f"Error en fila {row_idx}: {exc}") from exc

        self.stdout.write(self.style.SUCCESS(
            f"Importación completada. Categorías nuevas: {created_categories}, "
            f"Productos creados: {created_products}, Productos actualizados: {updated_products}"
        ))

    # ----------------------------
    # Helpers de lectura/mapeo XLS
    # ----------------------------
    def _read_headers(self, ws) -> Tuple[str, ...]:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = tuple((h or "").strip() for h in header_row)
        return headers

    def _row_to_dict(self, row: Tuple[Any, ...], headers: Tuple[str, ...]) -> Dict[str, Any]:
        return {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

    def _map_columns(self, headers: Tuple[str, ...]) -> Dict[str, str]:
        # Sinónimos aceptados por campo
        synonyms = {
            "category": {"categoria", "category", "cat", "familia"},
            "name": {"descripción corta"},
            "sku": {"referencia"},
            "price": {"precio", "price", "valor"},
            "stock": {"stock", "cantidad", "inventario", "existencias"},
            "description": {"descripción página web"},
            "image_url": {"imagen", "imagen_url", "image", "image_url", "foto", "foto_url"},
        }

        lower_headers = {h.lower(): h for h in headers if h}
        mapping: Dict[str, str] = {}
        for field, keys in synonyms.items():
            found = None
            for k in keys:
                if k in lower_headers:
                    found = lower_headers[k]
                    break
            if not found:
                # Campos obligatorios mínimos
                if field in {"category", "sku"}:
                    raise CommandError(f"No se encontró la columna obligatoria para '{field}'. Encabezados: {headers}")
            else:
                mapping[field] = found
        # El nombre puede faltar; usaremos SKU como nombre en ese caso
        if "name" not in mapping:
            mapping["name"] = mapping.get("sku", "sku")
        return mapping

    # ----------------------------
    # Helpers de parsing/conversión
    # ----------------------------
    def _parse_decimal(self, value: Any) -> Decimal | None:
        if value is None or value == "":
            return None
        if isinstance(value, (int, float, Decimal)):
            try:
                return Decimal(str(value))
            except Exception:
                return None
        if isinstance(value, str):
            txt = value.replace("$", "").replace(",", "").strip()
            try:
                return Decimal(txt)
            except InvalidOperation:
                return None
        return None

    # ----------------------------
    # Helpers de imágenes en disco/S3
    # ----------------------------
    def _load_images_for_product(
        self,
        *,
        product: Product,
        images_base_dir: str,
        category_name: str,
        sku: str,
        replace_existing: bool,
        limit: int,
    ) -> int:
        """
        Busca imágenes en disco bajo una estructura <base>/<categoria>/<SKU>/* y las carga:
        - La primera imagen se asigna a product.image_file
        - Todas las imágenes se agregan como VariantImage de la variante por defecto (name="")
        - Si replace_existing=True, elimina imágenes previas del producto/variante
        Retorna la cantidad de imágenes cargadas (0 si ninguna).
        """
        from products.models import ProductVariant, VariantImage  # import local para evitar ciclos en import

        image_paths = self._find_image_paths(images_base_dir, category_name, sku)
        if not image_paths:
            # Intentar usar default_image.* como fallback
            default_img = self._find_default_image_path(images_base_dir)
            if not default_img:
                return 0
            # Cargar solo la imagen por defecto
            image_paths = [default_img]
        if limit and limit > 0:
            image_paths = image_paths[:limit]

        # Obtener/crear variante por defecto
        variant, _ = ProductVariant.objects.get_or_create(product=product, name="")

        # Reemplazo: limpiar previas
        if replace_existing:
            try:
                # Limpiar product.image_file
                if getattr(product, "image_file", None) and product.image_file:
                    product.image_file.delete(save=False)
                product.image_url = ""  # evitar ambigüedad URL/archivo
                product.save()
            except Exception:
                # No interrumpir si falla la eliminación física
                pass
            # Borrar imágenes de la variante
            VariantImage.objects.filter(variant=variant).delete()

        # Evitar duplicados por nombre base
        existing_names = set()
        try:
            from django.db.models.functions import Reverse
            # No todos los backends soportan funciones específicas para basename; usamos Python abajo.
        except Exception:
            pass
        for vi in getattr(variant, "images", []).all():
            try:
                if vi.image_file and getattr(vi.image_file, "name", None):
                    existing_names.add(os.path.basename(vi.image_file.name))
            except Exception:
                continue

        loaded_count = 0
        main_assigned = False

        for idx, img_path in enumerate(image_paths):
            basename = os.path.basename(img_path)
            # Guardar imagen principal del producto con la primera encontrada
            if idx == 0:
                # Solo si no existe ya (cuando no se reemplaza)
                if not getattr(product, "image_file", None) or not product.image_file:
                    with open(img_path, "rb") as fh:
                        product.image_file.save(f"products/{sku}/{basename}", File(fh), save=True)
                main_assigned = True

            # Crear VariantImage si no hay duplicado de nombre
            if basename in existing_names and not replace_existing:
                continue

            from products.models import VariantImage as VI
            is_main_flag = (idx == 0)
            # Respetar unicidad de is_main: si ya existe un main y no estamos reemplazando, poner False
            if not replace_existing and is_main_flag:
                if VI.objects.filter(variant=variant, is_main=True).exists():
                    is_main_flag = False
            vi = VI(variant=variant, alt_text=product.name, sort_order=idx, is_main=is_main_flag)
            with open(img_path, "rb") as fh:
                vi.image_file.save(f"products/variants/{sku}/{basename}", File(fh), save=True)
            vi.save()
            loaded_count += 1

        # Garantizar que exista exactamente un main si cargamos algo
        if loaded_count > 0:
            mains = list(VariantImage.objects.filter(variant=variant, is_main=True).order_by("id"))
            if len(mains) == 0:
                first = VariantImage.objects.filter(variant=variant).order_by("sort_order", "id").first()
                if first:
                    first.is_main = True
                    first.save(update_fields=["is_main"])
            elif len(mains) > 1:
                # Dejar solo el primero como main
                for extra in mains[1:]:
                    extra.is_main = False
                    extra.save(update_fields=["is_main"])

        return loaded_count

    def _find_default_image_path(self, base_dir: str) -> str | None:
        """
        Busca un archivo llamado default_image con extensiones comunes en el directorio base:
        <base>/default_image.(jpg|jpeg|png|webp|gif)
        Retorna la ruta completa si existe.
        """
        allowed_ext = [".jpg", ".jpeg", ".png", ".webp", ".gif"]
        candidates = [os.path.join(base_dir, f"default_image{ext}") for ext in allowed_ext]
        for path in candidates:
            if os.path.isfile(path):
                return path
        # Búsqueda case-insensitive alternativa
        try:
            for fname in os.listdir(base_dir):
                root, ext = os.path.splitext(fname)
                if ext.lower() in allowed_ext and root.lower() == "default_image":
                    full = os.path.join(base_dir, fname)
                    if os.path.isfile(full):
                        return full
        except Exception:
            pass
        return None

    def _find_image_paths(self, base_dir: str, category_name: str, sku: str) -> list[str]:
        """
        Devuelve rutas de imagen bajo:
        - <base>/<categoria>/<SKU>/*.(jpg|jpeg|png|webp|gif)
        Si no existe, intenta localizar un directorio llamado como el SKU en cualquier subcarpeta.
        """
        allowed_ext = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

        def _casefold_match(name: str, candidates: list[str]) -> str | None:
            name_cf = name.casefold()
            for c in candidates:
                if c.casefold() == name_cf:
                    return c
            return None

        # Intento directo: <base>/<categoria>/<SKU>/
        try:
            cat_entries = os.listdir(base_dir)
            cat_dir_name = _casefold_match(category_name, cat_entries)
            if cat_dir_name:
                cat_dir = os.path.join(base_dir, cat_dir_name)
                sku_entries = os.listdir(cat_dir)
                sku_dir_name = _casefold_match(sku, sku_entries)
                if sku_dir_name:
                    sku_dir = os.path.join(cat_dir, sku_dir_name)
                    files = [
                        os.path.join(sku_dir, f)
                        for f in os.listdir(sku_dir)
                        if os.path.isfile(os.path.join(sku_dir, f))
                        and os.path.splitext(f)[1].lower() in allowed_ext
                    ]
                    files.sort()
                    return files
        except Exception:
            # Ignorar y probar búsqueda alternativa
            pass

        # Búsqueda alternativa: localizar directorio por SKU en cualquier subcarpeta (costo O(n))
        for dirpath, dirnames, filenames in os.walk(base_dir):
            # Coincidencia exacta (case-insensitive) del nombre del subdirectorio con el SKU
            for dn in dirnames:
                if dn.casefold() == sku.casefold():
                    sku_dir = os.path.join(dirpath, dn)
                    files = [
                        os.path.join(sku_dir, f)
                        for f in os.listdir(sku_dir)
                        if os.path.isfile(os.path.join(sku_dir, f))
                        and os.path.splitext(f)[1].lower() in allowed_ext
                    ]
                    files.sort()
                    return files
        return []
    def _make_unique_product_slug(self, name: str, sku: str, exclude_id: int | None = None) -> str:
        base = slugify(name) or slugify(sku) or sku.lower()
        candidate = base
        counter = 2
        qs = Product.objects.all()
        if exclude_id:
            qs = qs.exclude(id=exclude_id)
        while qs.filter(slug=candidate).exists():
            candidate = f"{base}-{counter}"
            counter += 1
        return candidate

    def _parse_int(self, value: Any) -> int | None:
        if value is None or value == "":
            return None
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            try:
                return int(value)
            except Exception:
                return None
        if isinstance(value, str):
            txt = value.strip()
            try:
                return int(txt)
            except Exception:
                return None
        return None


